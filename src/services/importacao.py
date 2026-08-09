from __future__ import annotations

import csv
import io
import json
import unicodedata
from collections.abc import Mapping
from datetime import date, datetime
from pathlib import Path
from typing import Any

from core.models import AnaliseLote

from .vencimento import calcular_dias_restantes, classificar_risco

EXTENSOES_ACEITAS = {".csv", ".xlsx", ".json", ".txt"}
TAMANHO_MAXIMO_BYTES = 5 * 1024 * 1024
ORIGEM_PADRAO = "IMPORTACAO"
LIMITE_CAMPO_TEXTO = 100

# Chaves canonicas do contrato persistido em AnaliseLote e aliases aceitos
# em arquivos de exportacao comuns (colunas podem vir em PT-BR, com acentos,
# espacos ou nomes comerciais).
ALIASES = {
    "codigo_produto": (
        "codigo_produto",
        "codigo",
        "codigoproduto",
        "produtocodigo",
        "codigodoproduto",
    ),
    "nome_produto": ("nome_produto", "produto", "descricao", "nome"),
    "quantidade_inicial": (
        "quantidade_inicial",
        "quantidade",
        "qtd",
        "saldo",
        "estoque",
    ),
    "lote": (
        "lote",
        "lotenro",
        "nlote",
        "numerolote",
        "numlote",
        "lote_numerico",
    ),
    "data_validade": (
        "data_validade",
        "validade",
        "vencimento",
        "datavencimento",
        "validadefinal",
        "datavalidade",
    ),
    "local": ("local", "deposito", "setor", "armazenagem"),
    "unidade": ("unidade", "volume", "embalagem"),
    "origem": ("origem", "fonte"),
}


def _normalizar_chave(valor: Any) -> str:
    texto = unicodedata.normalize("NFKD", str(valor))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto.strip().lower().replace(" ", "_").replace("-", "_")


def _normalizar_linha(registro: Mapping[str, Any]) -> dict[str, Any]:
    """Converte uma linha crua em contrato canonico, com aliases de coluna.

    Retorna o payload canonico quando os campos obrigatorios existem ou levanta
    ValueError com os erros de cada campo.
    """
    chaves = {_normalizar_chave(k): v for k, v in registro.items()}
    erros: dict[str, list[str]] = {}

    def extrair(campo: str) -> str:
        for alias in ALIASES[campo]:
            if alias in chaves:
                return str(chaves[alias]).strip()
        return ""

    codigo_produto = extrair("codigo_produto")
    nome_produto = extrair("nome_produto")
    lote = extrair("lote")
    quantidade_inicial = extrair("quantidade_inicial")
    data_validade = extrair("data_validade")
    local = extrair("local")
    unidade = extrair("unidade")
    origem = extrair("origem") or ORIGEM_PADRAO

    if not codigo_produto:
        erros["codigo_produto"] = ["Campo obrigatorio ausente ou vazio."]
    elif len(codigo_produto) > LIMITE_CAMPO_TEXTO:
        erros["codigo_produto"] = [
            f"Deve ter no maximo {LIMITE_CAMPO_TEXTO} caracteres."
        ]

    if not lote:
        erros["lote"] = ["Campo obrigatorio ausente ou vazio."]
    elif len(lote) > LIMITE_CAMPO_TEXTO:
        erros["lote"] = [f"Deve ter no maximo {LIMITE_CAMPO_TEXTO} caracteres."]

    quantidade_parsed = _converter_quantidade(quantidade_inicial)
    if quantidade_parsed is None:
        erros["quantidade_inicial"] = ["Campo obrigatorio ausente ou invalido."]
    elif quantidade_parsed < 0:
        erros["quantidade_inicial"] = ["Deve ser maior ou igual a zero."]

    if not data_validade:
        erros["data_validade"] = ["Campo obrigatorio ausente ou vazio."]
    else:
        data_parsed = _converter_data(data_validade)
        if data_parsed is None:
            erros["data_validade"] = [
                "Formato invalido. Use YYYY-MM-DD ou DD/MM/YYYY."
            ]

    if len(origem) > LIMITE_CAMPO_TEXTO:
        erros["origem"] = [f"Deve ter no maximo {LIMITE_CAMPO_TEXTO} caracteres."]
    if len(local) > LIMITE_CAMPO_TEXTO:
        erros["local"] = [f"Deve ter no maximo {LIMITE_CAMPO_TEXTO} caracteres."]
    if len(unidade) > 50:
        erros["unidade"] = ["Deve ter no maximo 50 caracteres."]

    if erros:
        raise ValueError(_resumir_erros(erros))

    return {
        "nome_produto": nome_produto,
        "codigo_produto": codigo_produto,
        "lote": lote,
        "quantidade_inicial": quantidade_parsed,
        "data_validade": data_parsed,
        "local": local,
        "unidade": unidade,
        "origem": origem,
    }


def _converter_quantidade(valor: Any):
    texto = str(valor).strip().replace(",", ".")
    if not texto:
        return None
    try:
        return float(texto)
    except (TypeError, ValueError):
        return None


def _converter_data(valor: Any) -> date | None:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    if not texto:
        return None
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def _resumir_erros(erros: dict[str, list[str]]) -> str:
    partes = []
    for campo, mensagens in erros.items():
        partes.append(f"{campo}: {'; '.join(mensagens)}")
    return " | ".join(partes)


def _parse_csv(conteudo: bytes, tipo: str) -> list[dict[str, Any]]:
    texto = conteudo.decode("utf-8-sig", errors="replace")
    if tipo == "txt":
        primeira = texto.splitlines()[0] if texto.splitlines() else ""
        if primeira.count("\t") > primeira.count(","):
            return list(csv.DictReader(io.StringIO(texto), delimiter="\t"))
    return list(csv.DictReader(io.StringIO(texto)))


def _parse_json(conteudo: bytes) -> list[dict[str, Any]]:
    try:
        dados = json.loads(conteudo.decode("utf-8-sig"))
    except (json.JSONDecodeError, UnicodeDecodeError) as exc:
        raise ValueError(f"JSON invalido: {exc}") from exc

    if isinstance(dados, list):
        linhas = dados
    elif isinstance(dados, dict):
        linhas = dados.get("lotes") or dados.get("registros") or dados.get("data")
    else:
        linhas = None

    if not isinstance(linhas, list):
        raise ValueError(
            "JSON deve ser uma lista de objetos ou um objeto com a chave 'lotes'."
        )

    registros: list[dict[str, Any]] = []
    for item in linhas:
        if not isinstance(item, dict):
            raise ValueError("Cada registro do JSON deve ser um objeto.")
        registros.append(item)
    return registros


def _parse_xlsx(conteudo: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "Dependencia openpyxl indisponivel no servidor para Excel."
        ) from exc

    workbook = load_workbook(
        io.BytesIO(conteudo),
        read_only=True,
        data_only=True,
    )
    planilha = workbook.active
    linhas = planilha.iter_rows(values_only=True)
    cabecalho = next(linhas, None)
    if not cabecalho:
        return []

    chaves = [_normalizar_chave(celula) if celula is not None else "" for celula in cabecalho]
    registros: list[dict[str, Any]] = []
    for valores in linhas:
        if valores is None or all(v is None for v in valores):
            continue
        registros.append(
            {chaves[i]: valores[i] for i in range(len(chaves)) if chaves[i]}
        )
    return registros


def _processar_registro(
    registro: dict[str, Any],
    hoje: date,
) -> tuple[AnaliseLote, bool]:
    payload = _normalizar_linha(registro)
    dias_restantes = calcular_dias_restantes(payload["data_validade"], hoje)
    classificacao = classificar_risco(payload["data_validade"], hoje)
    return AnaliseLote.objects.update_or_create(
        codigo_produto=payload["codigo_produto"],
        lote=payload["lote"],
        defaults={
            "data_validade": payload["data_validade"],
            "nome_produto": payload["nome_produto"],
            "quantidade_inicial": payload["quantidade_inicial"],
            "quantidade_atual": payload["quantidade_inicial"],
            "dias_restantes": dias_restantes,
            "classificacao": classificacao.value,
            "local": payload["local"],
            "unidade": payload["unidade"],
            "origem": payload["origem"],
        },
    )


def processar_importacao(
    nome: str,
    conteudo: bytes,
    hoje: date | None = None,
) -> dict[str, Any]:
    """Valida o arquivo, importa as linhas validas e resume erros por linha.

    Nao confia em cabecalhos ou tipos vindos do cliente: a validacao acontece
    aqui, no servidor, linha a linha. Nunca levanta para a view exceto por
    erros de arquivo (formato/limite); erros de linha sao coletados no resumo.
    """
    hoje = hoje or date.today()
    extensao = Path(nome).suffix.lower()

    if extensao not in EXTENSOES_ACEITAS:
        raise ValueError(
            "Formato de arquivo nao suportado. Use .csv, .xlsx, .json ou .txt. "
            "Arquivos .xls (Excel legado) devem ser convertidos para .xlsx."
        )
    if len(conteudo) > TAMANHO_MAXIMO_BYTES:
        raise ValueError(
            f"Arquivo excede o limite de {TAMANHO_MAXIMO_BYTES // (1024 * 1024)}MB."
        )

    if extensao == ".json":
        linhas = _parse_json(conteudo)
    elif extensao == ".xlsx":
        linhas = _parse_xlsx(conteudo)
    else:
        linhas = _parse_csv(conteudo, extensao.lstrip("."))

    resumo: dict[str, Any] = {
        "total": len(linhas),
        "importados": 0,
        "atualizados": 0,
        "invalidos": 0,
        "erros": [],
        "tipo_arquivo": extensao.lstrip("."),
    }

    for indice, registro in enumerate(linhas, start=1):
        try:
            _, criado = _processar_registro(registro, hoje)
        except ValueError as exc:
            resumo["invalidos"] += 1
            resumo["erros"].append(
                {"linha": indice, "erros": [str(exc)]}
            )
        except Exception as exc:  # noqa: BLE001 - mantem o resto do arquivo.
            resumo["invalidos"] += 1
            resumo["erros"].append(
                {"linha": indice, "erros": [f"Falha ao processar linha: {exc}"]}
            )
        else:
            if criado:
                resumo["importados"] += 1
            else:
                resumo["atualizados"] += 1

    return resumo
